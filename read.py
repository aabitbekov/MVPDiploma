import os
import numpy
import numpy as np
from pathlib import Path
import openpyxl
import pandas

def save_uploadedfile(uploadedfile):
    with open(os.path.join("excelFiles/{}".format(uploadedfile.name)), "wb") as f:
        f.write(uploadedfile.getbuffer())
    return "excelFiles/{}".format(uploadedfile.name)


def koef_prymyx_zatrat(main_matrix, val_price):
    koef_prymyx_zatrat = []
    row_koef_prymyx_zatrat = []
    index = 0
    for array in main_matrix:
        for element in array:
            try:
                row_koef_prymyx_zatrat.append(round(element/val_price[index], 4))
            except ZeroDivisionError:
                return 0
            index += 1
        koef_prymyx_zatrat.append(row_koef_prymyx_zatrat)
        row_koef_prymyx_zatrat = []
        index = 0
    return koef_prymyx_zatrat


def readDocument(path):
    xlsx_file = Path(path)
    wb_obj = openpyxl.load_workbook(xlsx_file)
    sheet = wb_obj.active
    return sheet


def readMainMatrix(sheet):
    # Read Matrix
    main_matrix, raw = [], []
    for row in sheet.iter_rows(2, sheet.max_row-5):
        for index in range(1,sheet.max_column-3):
            raw.append(row[index].value)
        main_matrix.append(raw)
        raw = []
    return main_matrix

def readByCol(sheet):
    # Read By Col
    col_names, itogo_by_row, dob_st, val_price_by_raw, trud, fondy = [], [], [], [], [], []
    for column in sheet.iter_cols(2, sheet.max_column):
        col_names.append(column[0].value)
        itogo_by_row.append(column[sheet.max_row - 5].value)
        dob_st.append(column[sheet.max_row - 4].value)
        val_price_by_raw.append(column[sheet.max_row - 3].value)
        trud.append(column[sheet.max_row - 2].value)
        fondy.append(column[sheet.max_row - 1].value)
    return col_names, itogo_by_row, dob_st, val_price_by_raw, trud, fondy

def readByRow(sheet):
    # Read By Row
    row_names, itogo_by_col, end_pruducts, val_price = [], [], [], []
    for row in sheet.iter_rows(2, sheet.max_row):
        row_names.append(row[0].value)
        itogo_by_col.append(row[sheet.max_column - 3].value)
        end_pruducts.append(row[sheet.max_column - 2].value)
        val_price.append(row[sheet.max_column - 1].value)
    return row_names, itogo_by_col, end_pruducts, val_price


def getInverseMatrix(main_matrix):
    inverse_matrix, identity_matrix = [], np.identity(len(main_matrix))
    b = numpy.linalg.inv(np.subtract(identity_matrix, main_matrix))
    for i in range(len(b)):
        for j in range(len(b)):
            b[i][j] = round(b[i][j], 3)

    return b


def getForWriter(main_matrix, itogo_by_col, end_pruducts, val_price, itogo_by_row, dob_st, val_price_by_raw, trud, fondy, col_names, row_names):
    i = 0
    for matrix in main_matrix:
        matrix.append(itogo_by_col[i])
        matrix.append(end_pruducts[i])
        matrix.append(val_price[i])
        i += 1
    main_matrix.append(itogo_by_row)
    main_matrix.append(dob_st)
    main_matrix.append(val_price_by_raw)
    main_matrix.append(trud)
    main_matrix.append(fondy)
    main_matrix = pandas.DataFrame(main_matrix)
    main_matrix.columns = col_names
    main_matrix.index = row_names
    return main_matrix


def readBigMainMatrix(sheet):
    main_matrix, raw , end_pruducts = [], [], []
    for row in sheet.iter_rows(0, sheet.max_row):
        for index in range(0, sheet.max_column-1):
             raw.append(row[index].value)
        main_matrix.append(raw)
        raw = []
    raw = 0
    for row in sheet.iter_rows(0, sheet.max_row):
        for index in range(sheet.max_column-1, sheet.max_column):
             raw = (row[index].value)
        end_pruducts.append(raw)
        raw = 0
    return main_matrix, end_pruducts





